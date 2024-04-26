from math import ceil
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from graphbrain_semsim import logger, RNG_SEED
from graphbrain_semsim.datasets.models import LemmaDataset

from graphbrain_semsim.datasets.config import CONFLICTS_ANNOTATION_LABELS

N_HEADER_ROWS: int = 3  # number of rows for the header


def make_dataset_table(
        lemma_dataset: LemmaDataset,
        output_path: Path,
        annotators: list[str] = None,
        divided_for_annotators: bool = False,
):
    logger.info(f"Making dataset table for '{lemma_dataset.id}'...")

    # Prepare columns for DataFrame
    idxes: list[int] = []
    lemmas: list[str] = []
    edge_texts: list[str] = []
    edges: list[str] = []

    # Iterate through lemma_matches and populate the columns
    for lemma, lemma_matches in lemma_dataset.lemma_matches.items():
        for lemma_match in lemma_matches:
            idxes.append(lemma_match.idx)
            lemmas.append(lemma_match.lemma)
            edge_texts.append(lemma_match.match.edge_text)
            edges.append(str(lemma_match.match.edge))

    # Create DataFrame
    df = pd.DataFrame({
        'Index': idxes,
        'Lemma': lemmas,
        'Edge Text': edge_texts,
        'Is Conflict?': [''] * lemma_dataset.n_samples,  # empty column for manual annotation
        'Edge': edges,
        'Bad Parse?': [''] * lemma_dataset.n_samples  # empty column for manual annotation
    })

    # Shuffle the entire DataFrame
    df = df.sample(frac=1, random_state=RNG_SEED).reset_index(drop=True)

    # Prepare Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write DataFrame to Excel starting from offset n_header_rows
        df.to_excel(writer, startrow=N_HEADER_ROWS, index=False, sheet_name='Dataset')

        # Access the created sheet and write the header (dataset's id)
        # and the annotation labels to Excel
        worksheet = writer.sheets['Dataset']
        worksheet.cell(row=1, column=1, value=f"Dataset id: {lemma_dataset.id}")
        worksheet.cell(row=2, column=1, value=f"Annotation labels: {ANNOTATION_LABELS}")

        if annotators:
            add_annotation_functionality(
                df,
                lemma_dataset.n_samples,
                N_HEADER_ROWS,
                writer,
                worksheet,
                annotators,
                divided_for_annotators,
                ANNOTATION_LABELS
            )


def add_annotation_functionality(
        df: pd.DataFrame,
        n_samples: int,
        n_header_rows: int,
        writer: pd.ExcelWriter,
        worksheet: Worksheet,
        annotators: list[str],
        divided_for_annotators: bool = False,
        annotation_labels: dict[str, int] = None,
):
    # Dividing the dataset among annotators if divided_for_annotators is True
    if divided_for_annotators:
        slice_size = ceil(len(df) / len(annotators))
        slices = [df.iloc[i:i + slice_size] for i in range(0, len(df), slice_size)]
    else:
        slices = [df for _ in annotators]

    # Write the slices to separate sheets for each annotator
    for name, annotator_df in zip(annotators, slices):
        annotator_df.to_excel(writer, index=False, sheet_name=name)

    first_annotator_col: int = len(df.columns) + 1  # column index for the first annotator
    last_annotator_col: int = first_annotator_col + len(annotators) - 1

    # Adding aggregated columns for annotators in the main sheet
    for col, name in enumerate(annotators, start=first_annotator_col):
        worksheet.cell(row=n_header_rows + 1, column=col, value=name)
        # Using Excel formula to copy the annotation from the annotator's sheet based on 'Index' column
        for row, idx in enumerate(df['Index'], start=n_header_rows + 2):
            formula = (
                f"=IF(ISNUMBER(MATCH({idx},'{name}'!$A:$A,0)), IF(OR("
                f"INDEX('{name}'!$D:$D,MATCH({idx},'{name}'!$A:$A,0))={annotation_labels['conflict']}, "
                f"INDEX('{name}'!$D:$D,MATCH({idx},'{name}'!$A:$A,0))={annotation_labels['no conflict']}), "
                f"INDEX('{name}'!$D:$D,MATCH({idx},'{name}'!$A:$A,0)), "
                f"0), 0)"  # return a "0" when there's no annotation
            )
            worksheet.cell(row=row, column=col, value=formula)

    # Adding formula to "Is Conflict?" column based on annotator's columns
    conflict_col = df.columns.get_loc("Is Conflict?") + 1
    for row in range(n_header_rows + 2, n_header_rows + n_samples + 1):
        conflict_formula = (
            # Check if all annotator cells for the row are zeros (indicating blanks). If all are zeros, return blank.
            f"=IF(SUM({get_column_letter(first_annotator_col)}{row}:{get_column_letter(last_annotator_col)}{row})=0, "
            f"\"\", "

            # Next, check if any of the annotator cells have a "conflict" annotation.
            f"IF(COUNTIF({get_column_letter(first_annotator_col)}{row}:{get_column_letter(last_annotator_col)}{row}, "
            f"\"{annotation_labels['conflict']}\") > 0, "

            # Then, within the above condition, check if there's any "no conflict" annotation. 
            # If there's both, return "disagree".
            f"IF(COUNTIF({get_column_letter(first_annotator_col)}{row}:{get_column_letter(last_annotator_col)}{row}, "
            f"\"{annotation_labels['no conflict']}\") > 0, "
            f"\"{annotation_labels['disagree']}\", "
            f"\"{annotation_labels['conflict']}\") ,"

            # If there wasn't any "conflict", check for "no conflict".
            f"IF(COUNTIF({get_column_letter(first_annotator_col)}{row}:{get_column_letter(last_annotator_col)}{row}, "
            f"\"{annotation_labels['no conflict']}\") > 0, "
            f"\"{annotation_labels['no conflict']}\", \"\")))"
        )

        worksheet.cell(row=row, column=conflict_col, value=conflict_formula)